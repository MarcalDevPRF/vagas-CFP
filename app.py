"""
app.py  ─  API de Lotação CFP  (Render)
========================================
Recebe dados do Google Apps Script e processa a alocação de lotação.

Colunas esperadas por arquivo
──────────────────────────────
alunos.csv   : inscricao_aluno, nome_aluno, cpf_aluno, concorrencia_aluno,
               situacao_aluno, email_aluno, pontuacao, data_nascimento
               (+ coluna 'classificacao' ou 'posicao_final' para ordenação)

vagas.csv    : unidade, vagas

respostas.csv: timestamp, tipo, papel, Inscrição_aluno, nome_aluno,
               email_aluno, pontos_aluno, concorrencia_aluno, situacao_aluno,
               cpf_aluno, telefone_aluno, acom_conjuge, matricula_conjuge,
               nome_conjuge, concorrencia_conjuge, situacao_conjuge,
               opcao_1 … opcao_28, selfie_file_id, assinatura_file_id,
               integridade_hash, protocolo

Endpoints
─────────
GET  /health               → liveness probe (Render)
POST /classificar          → recebe os 3 CSVs, processa, retorna JSON
POST /classificar?formato=pdf → idem, devolve PDF binário
GET  /resultado/csv        → baixa último resultado_lotacao.csv
GET  /resultado/xlsx       → baixa último resultado_lotacao.xlsx
GET  /saldo                → saldo final de vagas da última rodada
GET  /analise              → análise estatística da última rodada
"""

import io
import os
import json
import traceback
from pathlib import Path
from datetime import datetime

import pandas as pd
from flask import Flask, jsonify, request, send_file, abort
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Configuração
# ─────────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024   # 100 MB

WORKDIR = Path(os.environ.get("LOTACAO_WORKDIR", "/tmp/lotacao"))
WORKDIR.mkdir(parents=True, exist_ok=True)

SAIDA_CSV  = WORKDIR / "resultado_lotacao.csv"
SAIDA_XLSX = WORKDIR / "resultado_lotacao.xlsx"
META_JSON  = WORKDIR / "ultima_rodada.json"


# ═════════════════════════════════════════════════════════════════════════════
# LEITURA E NORMALIZAÇÃO DOS CSVs
# ═════════════════════════════════════════════════════════════════════════════

def _csv_to_df(file_storage) -> pd.DataFrame:
    """Lê FileStorage (ou bytes) para DataFrame."""
    raw = file_storage.read() if hasattr(file_storage, "read") else file_storage
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
        except Exception:
            continue
    raise ValueError("Não foi possível decodificar o CSV.")


def _col(df: pd.DataFrame, *nomes) -> str:
    """Retorna o primeiro nome de coluna encontrado no DataFrame (case-insensitive)."""
    mapa = {c.lower().strip(): c for c in df.columns}
    for n in nomes:
        if n.lower() in mapa:
            return mapa[n.lower()]
    return None


def carregar_alunos(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza o DataFrame de alunos. A posicao_final será calculada
    pela função montar_classificacao_final() após a normalização."""
    df.columns = df.columns.str.strip()

    # Inscrição
    col_insc = _col(df, "inscricao_aluno", "inscrição_aluno", "inscricao", "inscrição")
    if not col_insc:
        raise ValueError("Coluna de inscrição não encontrada em alunos.csv")
    df = df.rename(columns={col_insc: "inscricao_aluno"})
    df["inscricao_aluno"] = df["inscricao_aluno"].astype(str).str.strip()

    # Pontuação
    col_pont = _col(df, "pontuacao", "pontos", "pontos_aluno")
    if col_pont and col_pont != "pontuacao":
        df["pontuacao"] = pd.to_numeric(df[col_pont], errors="coerce").fillna(0)
    elif "pontuacao" not in df.columns:
        df["pontuacao"] = 0
    else:
        df["pontuacao"] = pd.to_numeric(df["pontuacao"], errors="coerce").fillna(0)

    # Data de nascimento → para desempate (maior idade = data menor = melhor)
    col_nasc = _col(df, "data_nascimento", "nascimento", "dt_nascimento")
    if col_nasc:
        df["data_nascimento"] = pd.to_datetime(
            df[col_nasc], dayfirst=True, errors="coerce"
        )
    else:
        df["data_nascimento"] = pd.NaT

    # Situação
    col_sit = _col(df, "situacao_aluno", "situacao", "situação")
    df["situacao_aluno"] = (df[col_sit].astype(str).str.strip().str.upper()
                            if col_sit else "REGULAR")

    # Nome
    col_nome = _col(df, "nome_aluno", "nome")
    df["nome_aluno"] = df[col_nome].astype(str).str.strip() if col_nome else "SEM NOME"

    # Concorrência
    col_conc = _col(df, "concorrencia_aluno", "concorrencia", "concorrência")
    df["concorrencia_aluno"] = (df[col_conc].astype(str).str.strip().str.upper()
                                if col_conc else "AMPLA")

    # Motivo chamada (opcional)
    col_mot = _col(df, "motivo_chamada")
    df["motivo_chamada"] = df[col_mot].astype(str).str.strip() if col_mot else ""

    return df


# ═════════════════════════════════════════════════════════════════════════════
# MONTAGEM DA CLASSIFICAÇÃO FINAL  (alternância + proporcionalidade)
# ═════════════════════════════════════════════════════════════════════════════

def montar_classificacao_final(df: pd.DataFrame) -> pd.DataFrame:
    """
    Monta a lista final de classificação com alternância e proporcionalidade.

    REGULARES:
      - 3 listas separadas por concorrência, ordenadas por nota DESC
        e desempate por data_nascimento ASC (mais velho = prioridade)
      - Relógio de posições reservadas:
          Pos 3, 8, 13, 18, 23... → COTA_NEGRO  (de 5 em 5, a partir de 3)
          Pos 5, 21, 41, 61, 81...→ COTA_PCD    (5 fixo, depois de 20 em 20)
          Demais → AMPLA
      - Se a cota reservada estiver esgotada → AMPLA preenche
      - Continua até todas as listas se esgotarem

    SUBJUDICE (vaga espelho):
      - Cada SUBJUDICE recebe a mesma posicao_final do REGULAR
        imediatamente de nota igual ou superior mais próxima
      - Nunca desloca a numeração ordinal dos REGULARES
    """
    # ── Separar REGULARES e SUBJUDICE ────────────────────────────────────────
    regulares  = df[df["situacao_aluno"] == "REGULAR"].copy()
    subjudices = df[df["situacao_aluno"] == "SUBJUDICE"].copy()

    # ── Ordenar cada lista: nota DESC, nascimento ASC ────────────────────────
    def ordenar_lista(sub_df):
        sub_df = sub_df.copy()
        sub_df["_nasc_sort"] = sub_df["data_nascimento"].apply(
            lambda d: d if pd.notna(d) else pd.Timestamp("2099-01-01")
        )
        return sub_df.sort_values(
            ["pontuacao", "_nasc_sort"], ascending=[False, True]
        ).reset_index(drop=True)

    listas = {
        "AMPLA":      ordenar_lista(regulares[regulares["concorrencia_aluno"] == "AMPLA"]),
        "COTA_NEGRO": ordenar_lista(regulares[regulares["concorrencia_aluno"] == "COTA_NEGRO"]),
        "COTA_PCD":   ordenar_lista(regulares[regulares["concorrencia_aluno"] == "COTA_PCD"]),
    }
    ptr = {"AMPLA": 0, "COTA_NEGRO": 0, "COTA_PCD": 0}

    # ── Relógio: tipo de cada posição ─────────────────────────────────────────
    def tipo_posicao(pos):
        # PcD: 5, 21, 41, 61... (5 fixo + de 20 em 20 a partir de 21)
        if pos == 5 or (pos >= 21 and (pos - 21) % 20 == 0):
            return "COTA_PCD"
        # Negro: 3, 8, 13, 18, 23... (de 5 em 5 a partir de 3)
        if (pos - 3) >= 0 and (pos - 3) % 5 == 0:
            return "COTA_NEGRO"
        return "AMPLA"

    # ── Montar fila de REGULARES ──────────────────────────────────────────────
    fila_regular = []
    pos_ordinal  = 0

    # Continua enquanto houver candidatos em QUALQUER lista
    while any(ptr[k] < len(listas[k]) for k in ptr):
        pos_ordinal += 1
        tipo_ideal = tipo_posicao(pos_ordinal)

        # Tenta o tipo ideal; se esgotado, tenta AMPLA; se AMPLA esgotada,
        # tenta qualquer cota restante
        if ptr[tipo_ideal] < len(listas[tipo_ideal]):
            tipo = tipo_ideal
        elif ptr["AMPLA"] < len(listas["AMPLA"]):
            tipo = "AMPLA"
        else:
            # Busca qualquer lista não esgotada
            tipo = next((k for k in ["COTA_NEGRO", "COTA_PCD"] if ptr[k] < len(listas[k])), None)
            if tipo is None:
                break

        cand = listas[tipo].iloc[ptr[tipo]].to_dict()
        ptr[tipo] += 1
        cand["posicao_final"] = pos_ordinal
        cand["tipo_vaga"]     = tipo
        fila_regular.append(cand)

    # ── Inserir SUBJUDICE (vaga espelho) ──────────────────────────────────────
    resultado_final = list(fila_regular)  # começa com todos os regulares

    for _, subj in subjudices.iterrows():
        row       = subj.to_dict()
        nota_subj = float(row.get("pontuacao", 0) or 0)

        # Encontra o REGULAR com nota igual ou imediatamente superior (mais próxima acima)
        # → SUBJUDICE recebe a mesma posicao_final desse REGULAR
        candidatos_acima = [
            r for r in fila_regular
            if float(r.get("pontuacao", 0) or 0) >= nota_subj
        ]
        if candidatos_acima:
            # O de menor nota entre os que estão acima = mais próximo abaixo do subj
            espelho = min(candidatos_acima, key=lambda r: float(r.get("pontuacao", 0) or 0))
            row["posicao_final"] = espelho["posicao_final"]
        else:
            # Nota maior que todos os regulares → posição 1
            row["posicao_final"] = 1

        row["tipo_vaga"] = row.get("concorrencia_aluno", "AMPLA")
        resultado_final.append(row)

    df_final = pd.DataFrame(resultado_final)

    if df_final.empty:
        return df_final

    # ── Ordenação final:
    #    1. posicao_final ASC
    #    2. REGULAR antes de SUBJUDICE na mesma posição
    #    3. nota DESC (para desempate entre SUBJUDICEs na mesma posição)
    df_final["_ord_sit"] = df_final["situacao_aluno"].apply(
        lambda s: 0 if s == "REGULAR" else 1
    )
    df_final = df_final.sort_values(
        ["posicao_final", "_ord_sit", "pontuacao"],
        ascending=[True, True, False]
    ).reset_index(drop=True)

    # Limpar colunas auxiliares
    df_final = df_final.drop(
        columns=[c for c in ["_nasc_sort", "_ord_sit"] if c in df_final.columns],
        errors="ignore"
    )

    return df_final


def carregar_vagas(df: pd.DataFrame) -> tuple:
    """Normaliza DataFrame de vagas. Retorna (saldo_vagas, vagas_originais)."""
    df.columns = df.columns.str.strip()
    col_unid = _col(df, "unidade", "nome_unidade", "unidade_nome")
    col_vaga = _col(df, "vagas", "qtd_vagas", "quantidade")
    if not col_unid or not col_vaga:
        raise ValueError("vagas.csv precisa ter colunas 'unidade' e 'vagas'")
    df["_unidade"] = df[col_unid].astype(str).str.strip().str.upper()
    df["_vagas"]   = pd.to_numeric(df[col_vaga], errors="coerce").fillna(0).astype(int)
    saldo = dict(zip(df["_unidade"], df["_vagas"]))
    return dict(saldo), dict(saldo)


def carregar_respostas(df: pd.DataFrame) -> tuple:
    """Normaliza DataFrame de respostas/escolhas. Retorna (df, opcao_cols)."""
    df.columns = df.columns.str.strip()

    # Inscrição
    col_insc = _col(df, "inscricao-aluno", "inscrição-aluno", "inscricao_aluno", "inscrição_aluno", "inscrição", "inscricao")
    if not col_insc:
        raise ValueError("Coluna de inscrição não encontrada em respostas.csv")
    df = df.rename(columns={col_insc: "Inscrição"})
    df["Inscrição"] = df["Inscrição"].astype(str).str.strip()

    # Filtrar apenas candidatos (excluir acompanhantes)
    col_papel = _col(df, "papel")
    if col_papel:
        df = df[df[col_papel].astype(str).str.strip().str.lower() != "acompanhante"].copy()

    # Manter apenas última resposta por inscrição (por timestamp)
    col_ts = _col(df, "timestamp")
    if col_ts:
        df[col_ts] = pd.to_datetime(df[col_ts], dayfirst=True, errors="coerce")
        df = (df.sort_values(col_ts)
                .groupby("Inscrição", as_index=False)
                .last()
                .reset_index(drop=True))

    # Colunas de opção (opcao_1 … opcao_N)
    opcao_cols = sorted(
        [c for c in df.columns if c.lower().startswith("opcao_")],
        key=lambda x: int(x.lower().split("_")[1])
    )
    for col in opcao_cols:
        df[col] = (df[col].astype(str).str.strip().str.upper()
                   .replace({"NAN": "", "NONE": ""}))

    # acom_conjuge
    col_acom = _col(df, "acom_conjuge")
    df["acom_conjuge"] = (df[col_acom].astype(str).str.strip().str.upper()
                          if col_acom else "NÃO")

    # situacao_conjuge
    col_sit_conj = _col(df, "situacao_conjuge")
    df["situacao_conjuge"] = (df[col_sit_conj].astype(str).str.strip().str.upper()
                              if col_sit_conj else "")

    # matricula_conjuge
    col_mat = _col(df, "matricula_conjuge")
    df["matricula_conjuge"] = df[col_mat].astype(str).str.strip() if col_mat else ""

    return df, opcao_cols


# ═════════════════════════════════════════════════════════════════════════════
# LÓGICA DE ALOCAÇÃO  (Regras R1–R8)
# ═════════════════════════════════════════════════════════════════════════════

def get_opcoes(row_escolha, opcao_cols):
    return [row_escolha[col] for col in opcao_cols
            if str(row_escolha.get(col, "")).strip()]


def _resultado(posicao, inscricao, nome, situacao,
               status, unidade, vagas_consumidas=0, obs=""):
    return {
        "posicao_final":    posicao,
        "inscricao_aluno":  inscricao,
        "nome_aluno":       nome,
        "situacao_aluno":   situacao,
        "unidade_alocada":  unidade or "",
        "status":           status,
        "vagas_consumidas": vagas_consumidas,
        "observacao":       obs,
    }


def alocar_candidato(cand, escolha, saldo_vagas, opcao_cols):
    inscricao    = str(cand["inscricao_aluno"])
    nome         = cand["nome_aluno"]
    situacao     = cand["situacao_aluno"]
    posicao      = cand["posicao_final"]
    consome_vaga = (situacao == "REGULAR")              # R5

    if escolha is None:
        return _resultado(posicao, inscricao, nome, situacao,
                          status="SEM_ESCOLHA", unidade=None,
                          obs="Candidato não encontrado nas respostas")

    acom_conjuge    = str(escolha.get("acom_conjuge", "")).strip().upper() in ("SIM", "S", "YES")
    sit_conjuge     = str(escolha.get("situacao_conjuge", "")).strip().upper()
    conjuge_regular = sit_conjuge == "REGULAR"

    opcoes = get_opcoes(escolha, opcao_cols)
    if not opcoes:
        return _resultado(posicao, inscricao, nome, situacao,
                          status="SEM_OPCAO", unidade=None,
                          obs="Nenhuma opção preenchida")

    for unidade in opcoes:
        saldo_atual = saldo_vagas.get(unidade, 0)
        vagas_nec   = 1
        if acom_conjuge and consome_vaga:               # R3 / R4
            vagas_nec = 2 if conjuge_regular else 1

        if saldo_atual >= vagas_nec:
            if consome_vaga:                            # R2
                saldo_vagas[unidade] -= vagas_nec

            obs_parts = []
            if not consome_vaga:
                obs_parts.append("SUBJUDICE: não consumiu vaga")
            if acom_conjuge:
                obs_parts.append("cônjuge REGULAR alocado junto (−2 vagas)"
                                 if conjuge_regular else "cônjuge SUBJUDICE (−1 vaga)")

            return _resultado(posicao, inscricao, nome, situacao,
                              status="ALOCADO", unidade=unidade,
                              vagas_consumidas=vagas_nec if consome_vaga else 0,
                              obs="; ".join(obs_parts))

    return _resultado(posicao, inscricao, nome, situacao,  # R8
                      status="NAO_ALOCADO", unidade=None,
                      obs=f"Sem vaga nas {len(opcoes)} opções informadas")


def _norm_insc(val):
    """Normaliza inscrição: remove .0 de leitura float, faz strip."""
    s = str(val).strip()
    return s[:-2] if s.endswith(".0") else s


def processar_alocacao(df_alunos_raw, df_respostas_raw, df_vagas_raw):
    # 1. Normalizar alunos
    df_alunos_norm          = carregar_alunos(df_alunos_raw)
    # 2. Montar lista final de classificação (alternância + SUBJUDICE espelho)
    classif                 = montar_classificacao_final(df_alunos_norm)
    saldo_vagas, vagas_orig = carregar_vagas(df_vagas_raw)
    respostas, opcao_cols   = carregar_respostas(df_respostas_raw)

    # Normaliza inscrição dos dois lados para garantir o match
    classif["inscricao_aluno"] = classif["inscricao_aluno"].apply(_norm_insc)
    respostas["Inscricao_norm"] = respostas["Inscrição"].apply(_norm_insc)
    idx        = respostas.set_index("Inscricao_norm")
    resultados = []

    for _, cand in classif.iterrows():                  # R1 — ordem da classificação
        insc    = _norm_insc(cand["inscricao_aluno"])
        escolha = idx.loc[insc].to_dict() if insc in idx.index else None

        resultado = alocar_candidato(cand, escolha, saldo_vagas, opcao_cols)
        resultado["concorrencia_aluno"] = cand.get("concorrencia_aluno", "")
        resultado["pontuacao"]          = cand.get("pontuacao", 0)
        resultado["motivo_chamada"]     = cand.get("motivo_chamada", "")
        resultados.append(resultado)

    df = pd.DataFrame(resultados)[[
        "posicao_final", "inscricao_aluno", "nome_aluno",
        "concorrencia_aluno", "situacao_aluno", "pontuacao",
        "motivo_chamada", "unidade_alocada", "status",
        "vagas_consumidas", "observacao",
    ]]
    return df, saldo_vagas, vagas_orig


# ═════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO XLSX
# ═════════════════════════════════════════════════════════════════════════════

CORES_STATUS = {"ALOCADO": "D4EDDA", "NAO_ALOCADO": "F8D7DA",
                "SEM_ESCOLHA": "FFF3CD", "SEM_OPCAO": "FFE0B2"}
CORES_SUBJ   = {"ALOCADO": "B8DFC4", "NAO_ALOCADO": "F1AFBB"}
LARGURAS     = {
    "posicao_final": 10, "inscricao_aluno": 14, "nome_aluno": 38,
    "concorrencia_aluno": 16, "situacao_aluno": 14, "pontuacao": 10,
    "motivo_chamada": 18, "unidade_alocada": 30, "status": 16,
    "vagas_consumidas": 16, "observacao": 45,
}


def exportar_xlsx(df: pd.DataFrame, caminho: Path):
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lotação")
        ws = writer.sheets["Lotação"]

        fill_hdr = PatternFill("solid", fgColor="1F3864")
        font_hdr = Font(bold=True, color="FFFFFF", size=11)
        borda    = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for cell in ws[1]:
            cell.font = font_hdr; cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center")

        status_idx = list(df.columns).index("status")
        sit_idx    = list(df.columns).index("situacao_aluno")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = row[status_idx].value or ""
            sit    = row[sit_idx].value    or ""
            cor    = (CORES_SUBJ.get(status, "FFFFFF")
                      if sit == "SUBJUDICE" else CORES_STATUS.get(status, "FFFFFF"))
            fill = PatternFill("solid", fgColor=cor)
            for cell in row:
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(vertical="center")

        for i, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = LARGURAS.get(col, 14)
        ws.freeze_panes = "A2"


# ═════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DE PDF
# ═════════════════════════════════════════════════════════════════════════════

def gerar_pdf(df: pd.DataFrame, certame: str) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles   = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph(f"Resultado de Lotação — {certame}", styles["Title"]))
    elementos.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5*cm))

    colunas = ["Pos.", "Inscrição", "Nome", "Concorrência", "Situação",
               "Pontuação", "Unidade Alocada", "Status", "Observação"]
    campos  = ["posicao_final", "inscricao_aluno", "nome_aluno", "concorrencia_aluno",
               "situacao_aluno", "pontuacao", "unidade_alocada", "status", "observacao"]
    dados   = [colunas] + [[str(r.get(c, "")) for c in campos] for _, r in df.iterrows()]

    t = Table(dados, repeatRows=1)
    estilo = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1F3864")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    STATUS_COR = {
        "ALOCADO":     colors.HexColor("#D4EDDA"),
        "NAO_ALOCADO": colors.HexColor("#F8D7DA"),
        "SEM_ESCOLHA": colors.HexColor("#FFF3CD"),
        "SEM_OPCAO":   colors.HexColor("#FFE0B2"),
    }
    status_idx = campos.index("status")
    for i, row_data in enumerate(dados[1:], start=1):
        cor = STATUS_COR.get(row_data[status_idx])
        if cor:
            estilo.add("BACKGROUND", (0, i), (-1, i), cor)

    t.setStyle(estilo)
    elementos.append(t)
    doc.build(elementos)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# ANÁLISE ESTATÍSTICA
# ═════════════════════════════════════════════════════════════════════════════

def gerar_analise(df: pd.DataFrame, saldo_final: dict, vagas_orig: dict) -> dict:
    total       = len(df)
    alocados    = df[df["status"] == "ALOCADO"]
    dist_status = df["status"].value_counts().to_dict()

    ocup_unidade = []
    for unidade, vagas in sorted(vagas_orig.items()):
        consumidas = vagas - saldo_final.get(unidade, 0)
        ocup_unidade.append({
            "unidade":         unidade,
            "vagas_total":     vagas,
            "vagas_ocupadas":  consumidas,
            "vagas_restantes": saldo_final.get(unidade, 0),
            "ocupacao_pct":    round(consumidas / vagas * 100, 1) if vagas else 0,
        })

    com_escolha   = total - dist_status.get("SEM_ESCOLHA", 0) - dist_status.get("SEM_OPCAO", 0)
    taxa_alocacao = round(len(alocados) / com_escolha * 100, 1) if com_escolha else 0

    return {
        "resumo": {
            "total_candidatos":  total,
            "alocados":          len(alocados),
            "nao_alocados":      dist_status.get("NAO_ALOCADO", 0),
            "sem_escolha":       dist_status.get("SEM_ESCOLHA", 0),
            "sem_opcao":         dist_status.get("SEM_OPCAO", 0),
            "taxa_alocacao_pct": taxa_alocacao,
        },
        "por_status":      dist_status,
        "por_situacao":    df["situacao_aluno"].value_counts().to_dict(),
        "por_concorrencia":df["concorrencia_aluno"].value_counts().to_dict(),
        "ocupacao_unidades": ocup_unidade,
        "top10_unidades_alocadas": (
            alocados["unidade_alocada"].value_counts().head(10)
            .rename_axis("unidade").reset_index(name="alocados")
            .to_dict(orient="records")
        ),
    }


# ═════════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/health")
def health():
    ultimo = None
    if META_JSON.exists():
        try:
            ultimo = json.loads(META_JSON.read_text()).get("rodada_ts")
        except Exception:
            pass
    return jsonify({"status": "ok", "ts": datetime.utcnow().isoformat(),
                    "ultima_rodada": ultimo})


@app.post("/classificar")
def classificar():
    """
    Recebe multipart/form-data com:
      - alunos     → CSV da aba Alunos_<certame>
      - respostas  → CSV da aba respostas_atual_<certame>
      - vagas      → CSV da aba vagas_<certame>
      - certame    → nome do certame (campo texto)
    Retorna JSON ou PDF conforme ?formato=pdf.
    """
    formato = request.args.get("formato", "json").lower()

    erros = [f"Campo '{c}' ausente."
             for c in ("alunos", "respostas", "vagas")
             if c not in request.files]
    if erros:
        return jsonify({"ok": False, "erros": erros}), 400

    certame = request.form.get("certame", "CFP")

    try:
        df_alunos    = _csv_to_df(request.files["alunos"])
        df_respostas = _csv_to_df(request.files["respostas"])
        df_vagas     = _csv_to_df(request.files["vagas"])
    except Exception as exc:
        return jsonify({"ok": False, "erro": f"Erro ao ler CSVs: {exc}"}), 400

    try:
        df, saldo_final, vagas_orig = processar_alocacao(df_alunos, df_respostas, df_vagas)
    except Exception as exc:
        return jsonify({"ok": False, "erro": str(exc),
                        "trace": traceback.format_exc()}), 500

    # Salvar resultado
    df.to_csv(SAIDA_CSV, index=False, encoding="utf-8-sig")
    exportar_xlsx(df, SAIDA_XLSX)

    analise   = gerar_analise(df, saldo_final, vagas_orig)
    rodada_ts = datetime.utcnow().isoformat()

    META_JSON.write_text(json.dumps({
        "rodada_ts": rodada_ts, "certame": certame,
        "analise": analise, "saldo_final": saldo_final,
    }, ensure_ascii=False, indent=2))

    # PDF
    if formato == "pdf":
        try:
            pdf_bytes = gerar_pdf(df, certame)
            return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                             as_attachment=False,
                             download_name=f"resultado_{certame}.pdf")
        except Exception as exc:
            return jsonify({"ok": False, "erro": f"Erro ao gerar PDF: {exc}"}), 500

    # JSON
    return jsonify({
        "ok":                    True,
        "certame":               certame,
        "rodada_ts":             rodada_ts,
        "analise":               analise,
        "primeiros_50_alocados": df[df["status"] == "ALOCADO"].head(50).to_dict(orient="records"),
        "download_csv":          "/resultado/csv",
        "download_xlsx":         "/resultado/xlsx",
    })


@app.get("/resultado/csv")
def resultado_csv():
    if not SAIDA_CSV.exists():
        abort(404, "Nenhuma rodada processada ainda.")
    return send_file(SAIDA_CSV, as_attachment=True,
                     download_name="resultado_lotacao.csv", mimetype="text/csv")


@app.get("/resultado/xlsx")
def resultado_xlsx():
    if not SAIDA_XLSX.exists():
        abort(404, "Nenhuma rodada processada ainda.")
    return send_file(SAIDA_XLSX, as_attachment=True,
                     download_name="resultado_lotacao.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/saldo")
def saldo():
    if not META_JSON.exists():
        abort(404, "Nenhuma rodada processada ainda.")
    meta = json.loads(META_JSON.read_text())
    return jsonify({"ok": True, "certame": meta.get("certame"),
                    "rodada_ts": meta["rodada_ts"], "saldo_final": meta["saldo_final"]})


@app.get("/analise")
def analise():
    if not META_JSON.exists():
        abort(404, "Nenhuma rodada processada ainda.")
    meta = json.loads(META_JSON.read_text())
    return jsonify({"ok": True, "certame": meta.get("certame"),
                    "rodada_ts": meta["rodada_ts"], **meta["analise"]})


@app.post("/debug")
def debug():
    """Devolve colunas, primeiras 3 linhas e diagnóstico de match de inscrições."""
    resultado = {}
    dfs = {}
    for campo in ("alunos", "respostas", "vagas"):
        if campo not in request.files:
            resultado[campo] = {"erro": "arquivo ausente"}
            continue
        try:
            df = _csv_to_df(request.files[campo])
            dfs[campo] = df
            resultado[campo] = {
                "colunas": list(df.columns),
                "linhas":  df.head(3).to_dict(orient="records"),
                "total":   len(df),
            }
        except Exception as e:
            resultado[campo] = {"erro": str(e)}

    # Diagnóstico de match entre alunos e respostas
    if "alunos" in dfs and "respostas" in dfs:
        df_a = dfs["alunos"]
        df_r = dfs["respostas"]

        # Encontrar coluna de inscrição nos alunos
        col_a = _col(df_a, "inscricao_aluno", "inscricao", "inscrição_aluno")
        # Encontrar coluna de inscrição nas respostas
        col_r = _col(df_r, "inscricao-aluno", "inscricao_aluno", "inscrição_aluno", "inscricao", "inscrição")

        inscs_alunos   = [str(v).strip() for v in df_a[col_a].tolist()[:10]] if col_a else []
        inscs_respostas = [str(v).strip() for v in df_r[col_r].tolist()[:10]] if col_r else []

        # Contar matches reais
        set_alunos    = set(str(v).strip() for v in df_a[col_a]) if col_a else set()
        set_respostas = set(str(v).strip() for v in df_r[col_r]) if col_r else set()
        matches = len(set_alunos & set_respostas)

        resultado["diagnostico_match"] = {
            "col_inscricao_alunos":    col_a,
            "col_inscricao_respostas": col_r,
            "primeiras_10_inscricoes_alunos":    inscs_alunos,
            "primeiras_10_inscricoes_respostas": inscs_respostas,
            "total_alunos":    len(set_alunos),
            "total_respostas": len(set_respostas),
            "matches_encontrados": matches,
            "sem_match_alunos": list(set_alunos - set_respostas)[:5],
            "sem_match_respostas": list(set_respostas - set_alunos)[:5],
        }

    return jsonify(resultado)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
